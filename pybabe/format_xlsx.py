
from base import MetaInfo, BabeBase

def valuenormalize(cell):
    "Build the row value out of a cell"
    if cell.number_format == '0': 
        return int(cell.internal_value)
    else: 
        return cell.internal_value

def read(format, instream, name, names, encoding, utf8_cleanup, **kwargs):
    from openpyxl import load_workbook
    wb = load_workbook(filename=instream, use_iterators=True)
    ws = wb.get_active_sheet()
    it = ws.iter_rows()
    if names: 
        yield MetaInfo(name=name, names = names)
    else:
        names_row = it.next()
        names = [cell.internal_value for cell in names_row]
        metainfo =  MetaInfo(name=name, names=names)
        yield metainfo
    for row in it: # it brings a new method: iter_rows()
        yield metainfo.t._make(map(valuenormalize, row))

def write(format, instream, outfile, encoding):
    from openpyxl import Workbook
    wb = Workbook(optimized_write = True)
    ws = wb.create_sheet()
    for k in instream:
        if isinstance(k, MetaInfo):
            metainfo = k
            ws.append(metainfo.names)
        else:
            ws.append(list(k))
    wb.save(outfile)

BabeBase.addPullPlugin('xlsx', ['xlsx'], read, need_seek=True)
BabeBase.addPushPlugin('xlsx', ['xlsx'], write)

